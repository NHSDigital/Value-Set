import time

from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Alignment
import setchks_app.setchks.setchk_definitions
from . import styling
from .termbrowser import termbrowser_hyperlink
 
def make_analysis_by_row_sheet(
    ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    output_OK_messages=None,
    analysis_by_outcome_row_numbers_map=None,
    supp_tabs_row_numbers_map=None,
    ): 

    # vsmt_style_wrap_top=NamedStyle(name="vsmt_style_wrap_top")
    # vsmt_style_wrap_top.alignment=Alignment(wrap_text=True, vertical='top')

    # vsmt_style_grey_row=NamedStyle(name="vsmt_style_grey_row")
    # vsmt_style_grey_row.fill=color_fills["grey"]
    # vsmt_style_grey_row.border=border

    setchks=setchks_app.setchks.setchk_definitions.setchks
    setchks_results=setchks_session.setchks_results

    row_analysis_row_numbers_map=[] # each entry in list corresponds 1:1 to a row in data file
                                # each such entry is a dict
                                #        keyed by:setchk code
                                #        value=row number in row_analysis_sheet that this function constructs 
                                # structure is used so that other sheets can link to the right row on this sheet 

    # simple header row (but only if data matrix had one; need to do this better)
    current_ws_row=0
    if setchks_session.table_has_header:
        header_row_cell_contents=[x.string for x in setchks_session.data_as_matrix[0]]
        # ws.append(["Row number", "Check", "Message"] + setchks_session.data_as_matrix[0]) # ** need to create better header row
        ws.append(["Row number", "Check", "Message","Row specific info","Link","Supp tab link"] + header_row_cell_contents) # ** need to create better header row
        current_ws_row+=1
    ws_append_time=0
    time_bigloop0=time.time()
    n_loops=0
    total_dt_inner_loop=0
    max_dt_inner_loop=0
    min_dt_inner_loop=9999999
    mixed_column=setchks_session.columns_info.mixed_column
    for i_data_row, data_row in enumerate(setchks_session.data_as_matrix[setchks_session.first_data_row:]):
        # data_row_cell_contents=[x.string for x in data_row]
        cid=setchks_session.marshalled_rows[i_data_row].C_Id
        cid_entered=setchks_session.marshalled_rows[i_data_row].C_Id_entered
        data_row_cell_contents=[]
        for i_col, cell_content in enumerate(data_row): 
            if i_col==mixed_column and cid is not None and cid_entered is not None: # don't hyperlink D_Id
                                                                                    # or non-SCTID
                                                                                    # hyperlink D_Id is confusing and the implied C_Id is given nearby
                data_row_cell_contents.append(
                    termbrowser_hyperlink(
                        sctid=cell_content.string, 
                        destination_sctid=cid,
                        )
                    )
            else:
                data_row_cell_contents.append(cell_content.string)

        
        something_was_output=False
        row_analysis_row_numbers_map.append({})
        current_row_map=row_analysis_row_numbers_map[-1] 
        for setchk_code in setchks_list_to_report:
            setchk_short_name=setchks[setchk_code].setchk_short_name
            setchk_results=setchks_results[setchk_code]
            if setchk_results.row_analysis!=[]:
                if setchk_code in supp_tabs_row_numbers_map: # i.e. there is a supp tab for this check
                    supp_tab_ws, supp_tab_mapping=supp_tabs_row_numbers_map[setchk_code]
                    # print(f"supp_tab_mapping:{supp_tab_mapping}")
                else:
                    supp_tab_ws=None        
                # data_row_cell_contents=[x.string for x in data_row]
                # ws.append([i_data_row+setchks_session.first_data_row+1, setchk_short_name, setchk_results.row_analysis[i_data_row]["Message"]]+data_row_cell_contents)
                outcome_codes_count={} # this is used to make sure that in the case where the same outcome_code
                                    # can occur more than once for the same row (e.g. where checking for unreccomended tl-hierarchies
                                    # and the concept is in more than one tl-hierarchy then) then the hyperlink fgoes to the
                                    # correct row of the "by outcome" table
                                    # this has been implemented but not thoroughly tested yet! 
                for check_item in setchk_results.row_analysis[i_data_row]:
                    if output_OK_messages or check_item.outcome_level not in ["INFO","DEBUG"]:
                        time_inner_0=time.time()
                        outcome_code=check_item.outcome_code
                        if outcome_code not in outcome_codes_count:
                            outcome_codes_count[outcome_code]=0
                        else:
                            outcome_codes_count[outcome_code]+=1
                        row_to_link_to=analysis_by_outcome_row_numbers_map[outcome_code][i_data_row][outcome_codes_count[outcome_code]]
                        message=f"{outcome_code}:{check_item.general_message}" 
                        row_specific_message=check_item.row_specific_message
                        if row_specific_message == "None":
                            row_specific_message=""
                        hyperlink_cell_contents=f'=HYPERLINK("#By_Outcome!B{row_to_link_to}","X")'
                        # print(f"i_data_row: {i_data_row} supp_tab_ws: {supp_tab_ws}")
                        if supp_tab_ws is not None:
                            # print(f"supp_tab_mapping:{supp_tab_mapping} {i_data_row} {supp_tab_mapping[i_data_row]}")    
                            if supp_tab_mapping[i_data_row] is not None:
                                row_to_link_to=supp_tab_mapping[i_data_row]
                                # print(f"row_to_link_to {i_data_row} {row_to_link_to}")
                                supp_tab_hyperlink_cell_contents=f'=HYPERLINK("#{supp_tab_ws.title}!A{row_to_link_to}","S")'
                            else:
                                supp_tab_hyperlink_cell_contents=""
                        else:
                            supp_tab_hyperlink_cell_contents=""

                        # print(f"MESSAGE_CCELL_CONTENTS:{message_cell_contents}")
                        # print(len(message_cell_contents))
                        ws_row_contents=[
                            i_data_row+setchks_session.first_data_row+1, 
                            setchk_short_name, 
                            message,
                            row_specific_message,
                            hyperlink_cell_contents,
                            supp_tab_hyperlink_cell_contents,
                            ] 
                        if not something_was_output: # only add the file data for the first outcome line
                            ws_row_contents+=data_row_cell_contents
                        time0=time.time()
                        ws.append(ws_row_contents)
                        current_ws_row+=1
                        dt=time.time()-time0
                        ws_append_time+=dt
                        n_loops+=1
                        # current_row_map[setchk_code]=ws.max_row
                        current_row_map[setchk_code]=current_ws_row
                        something_was_output=True
                        dt_inner_loop=time.time()-time_inner_0
                        total_dt_inner_loop+=dt_inner_loop
                        max_dt_inner_loop=max(max_dt_inner_loop, dt_inner_loop)
                        min_dt_inner_loop=min(min_dt_inner_loop, dt_inner_loop)
                        if dt_inner_loop>0.003:
                            print(setchk_code,i_data_row, dt_inner_loop)
        if something_was_output:
            ws.append(["----"]) 
            current_ws_row+=1
    print(f"Big loop time: {time.time()-time_bigloop0}")
    print(f"ws_append_time: {ws_append_time}")
    print(f"n_loops: {n_loops}")
    print(f"total_dt_inner_loop{total_dt_inner_loop}")
    print(f"max_dt_inner_loop{max_dt_inner_loop}")
    print(f"min_dt_inner_loop{min_dt_inner_loop}")
    # crude cell with setting
    cell_widths=[15,30,50,30,5,5,25,50] + [20]*10
    for i, width in enumerate(cell_widths):
        ws.column_dimensions[get_column_letter(i+1)].width=width     

    # example bit of formatting bling
    irow=0
    for row in ws.iter_rows():
        irow+=1
        divider_line=row[0].value=="----"
        if divider_line:
            ws.row_dimensions[irow].height = 3
        for cell in row:
            # cell.alignment=Alignment(wrap_text=True, vertical='top')
            # cell.style=vsmt_style_wrap_top
            # cell.alignment=cell.alignment.copy(wrap_text=True, vertical='top')
            if divider_line:
                cell.style=styling.vsmt_style_grey_row
                # cell.fill=color_fills["grey"]
                # cell.border = border
            else:
                strval=str(cell.value)
                if len(strval)>=16 and str(cell.value)[0:16]=='=HYPERLINK("http':
                    cell.style=styling.vsmt_style_wrap_top_hyperlink
                elif len(strval)>=6 and str(cell.value)[0:6]=='=HYPER':
                    cell.style=styling.vsmt_style_wrap_top_double_hyperlink
                else:
                    cell.style=styling.vsmt_style_wrap_top
            
    return row_analysis_row_numbers_map
