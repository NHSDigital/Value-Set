
from openpyxl.utils import get_column_letter

def make_supp_tab_sheets(
    wb=None,
    first_i_ws=None, 
    setchks_list_to_report=None,
    setchks_session=None,
    color_fills=None,
    border=None,
    ):

    supp_tabs_row_numbers_map={} # keyed by setchk code
                                 # values are tuples (ws, supp_tab_row_numbers_map) 
                                   
    i_ws=first_i_ws-1
    for setchk_code in setchks_list_to_report:
        setchk_results=setchks_session.setchks_results[setchk_code]
        if setchk_results.supp_tab_blocks is not None:
            i_ws+=1
            ws=wb.worksheets[i_ws]
            ws.title=f"{setchk_code}_supp_tab"
            supp_tab_row_numbers_map=make_one_supp_tab_sheet(
                ws=ws,
                setchk_code=setchk_code,
                setchk_results=setchk_results,
                color_fills=color_fills,
                border=border,
                )
            supp_tabs_row_numbers_map[setchk_code]=(ws, supp_tab_row_numbers_map)
    return i_ws, supp_tabs_row_numbers_map

def make_one_supp_tab_sheet(
    ws=None,
    setchk_code=None,
    setchk_results=None,
    color_fills=None,
    border=None,
    ):

    supp_tab_row_numbers_map=[]    # each entry in list corresponds 1:1 to a row in data file
                                    # each such entry is a row number in supp_tab that this function constructs 
                                    # so that other sheets can link to the right row on this sheet 
    headers_output=False
    banded_row=False
    for i_data_row, supp_tab_entries in enumerate(setchk_results.supp_tab_blocks):
        if supp_tab_entries not in [None, []]:
            banded_row=not banded_row
            first_row_of_block=True
            for supp_tab_row in supp_tab_entries:
                
                ##############################################
                #        output header row                   #
                # (delayed until have first SuppTabRow       #
                # object to get class level attributes from) #
                if not headers_output:
                    ws.append(supp_tab_row.headers) # this is a class level attribute. 
                                                    # Just pull it the first time fine a not "None" entry
                    for i, width in enumerate(supp_tab_row.cell_widths):
                        ws.column_dimensions[get_column_letter(i+1)].width=width
                    for cell in ws[ws.max_row]:
                        cell.alignment=cell.alignment.copy(wrap_text=True, horizontal="center")
                    headers_output=True
                #                                            #
                ##############################################
                ws.append(supp_tab_row.format_as_list())
                for cell in ws[ws.max_row]:
                    cell.alignment=cell.alignment.copy(wrap_text=True)
                    cell.border = border  
                    if banded_row:
                        cell.fill=color_fills["light_grey_band"]
                if first_row_of_block:
                    supp_tab_row_numbers_map.append(ws.max_row)
                    first_row_of_block=False
        else:
            supp_tab_row_numbers_map.append(None)
    return supp_tab_row_numbers_map